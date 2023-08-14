from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import struct
import pyaudio
import pyttsx3
import speech_recognition as sr
import time

wb = Workbook()
wb = load_workbook("sample.xlsx")
ws = wb.active

a="multiply a1 and a2"
ch ='y'
while(ch=='y'):
    ##break
    r = sr.Recognizer()
    with sr.Microphone() as source: 
        print("Listening ...")
        audio = r.listen(source)
        query = ''

        try:
            query = r.recognize_google(audio, language = 'en-IN', show_all=True)
            print(query['alternative'][0]['transcript'])
            a=  query['alternative'][0]['transcript'].lower()
            ch=input("Do you want to continue y/n")
            # if a.__contains__("stop"):
            #     break
        # finally:
        #     pass
        except Exception as e:
            print("Speak loudly") 

import spacy
from spacy.tokens import DocBin
from tqdm import tqdm
from spacy.matcher import Matcher

operation = []
cell = []

nlp = spacy.load("content\model-best") 
transcriptions = a
doc = nlp(transcriptions)
print(transcriptions)
print(doc.ents)
matcher = Matcher(nlp.vocab)
pattern = [{"ENT_TYPE": "OPERATION"}]
matcher.add("", [pattern]) 
matches = matcher(doc)
for match_id, start, end in matches:
    string_id = nlp.vocab.strings[match_id]  # Get string representation
    span = doc[start:end]  # The matched span
    if pattern[0]["ENT_TYPE"]=="OPERATION":
        operation.append(span.text)
        print(span.text)
print (operation)
# ###########################################################################
mat = Matcher(nlp.vocab)
patt = [{"ENT_TYPE": "CELL"}]
mat.add("", [patt]) 
match = mat(doc)
for match_id, start, end in match:
    string_id = nlp.vocab.strings[match_id]  # Get string representation
    span = doc[start:end]  # The matched span
    if patt[0]["ENT_TYPE"]=="CELL":
        cell.append(span.text.upper())
        print(span.text)
print("in cell")
print(cell)
if operation[0]=="add":
    ws[cell[2]].value=ws[cell[0]].value+ws[cell[1]].value
    wb.save("sample.xlsx")
if operation[0]=="subract":
    ws[cell[2]].value=ws[cell[0]].value-ws[cell[1]].value
    wb.save("sample.xlsx")
if operation[0]=="multiply":
    ws[cell[2]].value=ws[cell[0]].value*ws[cell[1]].value
    wb.save("sample.xlsx")
if operation[0]=="divide":
    ws[cell[2]].value=ws[cell[0]].value/ws[cell[1]].value
    wb.save("sample.xlsx")

    
#spacy.displacy.serve(doc, style='ent')